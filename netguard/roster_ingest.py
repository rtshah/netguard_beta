"""Deterministic Excel roster ingestion (Module 03 Part B)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .roster_schema import Roster, RosterPlan
from .standardize import standardize_address, standardize_name

# Canonical column aliases → internal field
_COLUMN_ALIASES: dict[str, str] = {
    "plan_id": "plan_id",
    "plan id": "plan_id",
    "planid": "plan_id",
    "plan_number": "plan_id",
    "plan number": "plan_id",
    "plan_name": "plan_name",
    "plan name": "plan_name",
    "planname": "plan_name",
    "name": "plan_name",
    "address": "address",
    "plan_address": "address",
    "formulary_id": "formulary_id",
    "formulary id": "formulary_id",
    "formularyid": "formulary_id",
    "lives": "lives",
    "enrollment": "lives",
    "member_lives": "lives",
    "alt_id": "alt_id",
    "alternate_id": "alt_id",
    "stem": "alt_id",  # demo generator stores formulary stem here
}


class RosterIngestError(Exception):
    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _map_headers(headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, h in enumerate(headers):
        key = _COLUMN_ALIASES.get(_norm_header(h))
        if key:
            mapping[idx] = key
    return mapping


def ingest_roster_excel(
    path: Path,
    *,
    payer_or_pbm: str | None = None,
    roster_period: str | None = None,
) -> Roster:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    sheet_title = ws.title or ""
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    errors: list[str] = []
    if not rows:
        raise RosterIngestError([f"{path}: empty workbook"])

    col_map = _map_headers(list(rows[0]))
    if "plan_id" not in col_map.values() or "plan_name" not in col_map.values():
        errors.append("roster must include plan_id and plan_name columns")
        raise RosterIngestError(errors)

    # Optional metadata from sheet name: "Express Scripts|2026-03"
    meta_payer = payer_or_pbm
    meta_period = roster_period
    if "|" in sheet_title:
        parts = sheet_title.split("|", 1)
        meta_payer = meta_payer or parts[0].strip()
        meta_period = meta_period or parts[1].strip()

    if not meta_payer:
        errors.append("payer_or_pbm is required (pass explicitly or encode in sheet title)")
    if not meta_period:
        errors.append("roster_period is required (YYYY-MM; pass explicitly or encode in sheet title)")
    if errors:
        raise RosterIngestError(errors)

    plans: list[RosterPlan] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        fields: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                fields[field] = row[idx]
        plan_id = str(fields.get("plan_id") or "").strip()
        plan_name = str(fields.get("plan_name") or "").strip()
        if not plan_id or not plan_name:
            errors.append(f"row {row_num}: missing plan_id or plan_name")
            continue
        address = fields.get("address")
        address_str = str(address).strip() if address is not None and str(address).strip() else None
        fid = fields.get("formulary_id")
        fid_str = str(fid).strip() if fid is not None and str(fid).strip() else None
        lives_raw = fields.get("lives")
        lives: float | None
        try:
            lives = float(lives_raw) if lives_raw is not None and str(lives_raw).strip() != "" else None
        except (TypeError, ValueError):
            errors.append(f"row {row_num}: lives is not numeric ({lives_raw!r})")
            lives = None
        alt_ids: dict[str, str] = {}
        if fields.get("alt_id") is not None and str(fields["alt_id"]).strip():
            alt_ids["alt"] = str(fields["alt_id"]).strip()

        plans.append(
            RosterPlan(
                plan_id=plan_id,
                plan_name=plan_name,
                plan_name_standardized=standardize_name(plan_name),
                address=address_str,
                address_standardized=standardize_address(address_str),
                formulary_id=fid_str,
                lives=lives,
                alt_ids=alt_ids,
            )
        )

    if errors:
        raise RosterIngestError(errors)
    if not plans:
        raise RosterIngestError([f"{path}: no plan rows found"])

    return Roster(
        payer_or_pbm=meta_payer or "",
        roster_period=meta_period or "",
        source_file=str(path),
        plans=plans,
    )
